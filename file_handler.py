# file_handler.py
import logging
import os
import csv

import numpy as np
from openpyxl import load_workbook  # 假设你用openpyxl处理Excel文件
from werkzeug.utils import secure_filename
from model import predict, scaler, pca


output_folder = os.path.join(os.getcwd(), 'output', 'predictions')  # 指定输出文件夹名称


def process_row_data(row_data):
    prediction = predict(row_data)
    return prediction


def process_uploaded_file(uploaded_file):
    # 检查文件是否有效并被上传
    logging.info('喵喵到这里啦~')
    if uploaded_file and uploaded_file.filename:
        # 确定临时存储路径
        file_path = secure_filename(uploaded_file.filename)
        uploaded_file.save(file_path)  # 存储到服务器的一个位置
        logging.info('进行到这里啦~')

        try:
            # 打开和处理Excel文件
            wb = load_workbook(filename=file_path)
            ws = wb.active

            data_list = []
            cnt = 0
            for row in ws.iter_rows():
                print(row)
                cnt += 1
                if cnt <= 1:continue

                row_data = [cell.value for cell in row]

                prediction = process_row_data(row_data)
                row_data.append(prediction[0])
                data_list.append(row_data)

            print(data_list)


            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
            output_csv_file = os.path.join(output_folder, 'predictions.csv')
            with open(output_csv_file, 'w', newline='') as f:
                writer = csv.writer(f)
                # 写入表头（假设为“Input”, “Prediction”）
                writer.writerow(['activity', 'attention', 'stars', 'technical_fork', 'participants',
                                 'new_contributors', 'inactive_contributors', 'bus_factor', 'issues_new',
                                 'issues_closed', 'issue_comments', 'issue_response_time',
                                 'issue_resolution_duration', 'issue_age', 'code_change_lines_add',
                                 'code_change_lines_remove', 'code_change_lines_sum', 'change_requests',
                                 'change_requests_accepted', 'change_requests_reviews', 'openrank'])
                for item in data_list:
                    writer.writerow(item)


        except Exception as e:
            print(f"处理文件时出错: {e}")

        os.remove(file_path)

        return True
    else:
        return False
